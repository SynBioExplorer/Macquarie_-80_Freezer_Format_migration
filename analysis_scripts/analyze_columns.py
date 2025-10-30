#!/usr/bin/env python3
import openpyxl
from openpyxl import load_workbook
import os
from collections import defaultdict
import json

# List of all Excel files to analyze
excel_files = [
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/BLANK TEMPLATE - PC2 Lab Sample Recording Template for -80 Freezers_May 2025.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/Freezer allocations Walk-in freezer PC2 Synbio.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80C freezer2 log_box level.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/Rack C2 -80oC freezer 3-Dayane-Costa-Williams.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/rack A4 -80C freezer 3_AfrinTalukder&VivianBonacker.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/Rack 2E -80oC freezer 3-Josh-Carla.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/Rack 2A_-80 Freezer 3_ FP.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/rack B2 -80C freezer 3_VictoriaBarja.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/rack A1 -80C freezer 3 .xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/Rack C3 -80C freezer3_JordiPerez.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/rack 2A -80C freezer 3.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/rack 2D -80C, freezer 3, 1st column SP.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer3 log/rack 2C -80C freezer 3_ArturSawiki.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer 1 log/-80 freezer 1_rack D1_EA&JP.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer 1 log/-80C freezer 1_rack D2_VictoriaBarja.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer 1 log/-80C Freezers_1_Rack_D3_KM.csv.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer 1 log/-80C Freezer1 Rack B2 DanielPascoe.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer 1 log/-80C freezer1 log_box level.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer 1 log/-80C Freezer1 Rack C1 ArdenL&NickY.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer 1 log/-80 Freezer1 log_tube level/RackB1 -80C Freezer 1.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf C/rack C6 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf C/rack C3 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf C/rack C5 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf C/rack C1 -80C freezer 2_LV.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf C/rack C4 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf D/2023-11-30 rack D1 -80C freezer 2 - backup.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf D/rack D6 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf D/rack D5 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf D/rack D4 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf D/rack D2 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf D/rack D1 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf E/rack E5 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf E/rack E6 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf E/rack E4 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf E/rack E2 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf E/rack E1 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf E/rack E3 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf B/rack B1 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf B/rack B2 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf B/rack B3 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf B/rack B6 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf B/rack B5 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf B/rack B4 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf A/rack A3 -80C freezer 2 (SC).xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf A/rack A4 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf A/rack A5 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf A/rack A2 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf A/rack A6 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf A/rack A1 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf D/rack D3 -80C freezer 2.xlsx",
    "/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting/older freezer log format/-80 freezer2 log/-80 freezer2 log_tube level/shelf C/rack C2 -80C freezer 2.xlsx",
]

def get_cell_value(cell):
    """Get the string value of a cell, handling None values"""
    if cell.value is None:
        return None
    return str(cell.value).strip()

def find_header_row(sheet, max_rows=20):
    """Find the row that contains column headers"""
    for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
        row = sheet[row_idx]
        # Count non-empty cells
        non_empty = sum(1 for cell in row if cell.value is not None and str(cell.value).strip())
        # If we have at least 3 non-empty cells, consider it a header row
        if non_empty >= 3:
            return row_idx
    return 1  # Default to first row

def extract_columns_from_file(file_path):
    """Extract column headers and sample data from an Excel file"""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        results = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Find the header row
            header_row_idx = find_header_row(sheet)

            # Extract headers
            headers = []
            header_row = sheet[header_row_idx]
            for cell in header_row:
                val = get_cell_value(cell)
                if val:
                    headers.append(val)
                elif headers:  # Stop at first empty cell after we've started collecting headers
                    break

            if not headers:
                continue

            # Extract sample data (next 3 rows)
            sample_data = []
            for row_idx in range(header_row_idx + 1, min(header_row_idx + 4, sheet.max_row + 1)):
                row_data = []
                row = sheet[row_idx]
                for idx, cell in enumerate(row[:len(headers)]):
                    val = get_cell_value(cell)
                    row_data.append(val if val else "")
                # Only include row if it has some data
                if any(row_data):
                    sample_data.append(row_data)

            results[sheet_name] = {
                'headers': headers,
                'sample_data': sample_data,
                'header_row': header_row_idx
            }

        wb.close()
        return results

    except Exception as e:
        return {'error': str(e)}

# Main analysis
print("=" * 100)
print("ANALYZING ALL EXCEL FILES IN OLDER FREEZER LOG FORMAT")
print("=" * 100)

# Track all unique columns and their frequency
all_columns = defaultdict(lambda: {'count': 0, 'files': [], 'example_file': None})
file_formats = {}

for file_path in excel_files:
    file_name = os.path.basename(file_path)
    print(f"\n{'=' * 100}")
    print(f"FILE: {file_name}")
    print(f"PATH: {file_path}")
    print('-' * 100)

    results = extract_columns_from_file(file_path)

    if 'error' in results:
        print(f"ERROR: {results['error']}")
        continue

    file_formats[file_name] = results

    for sheet_name, sheet_data in results.items():
        headers = sheet_data['headers']
        sample_data = sheet_data['sample_data']

        print(f"\nSHEET: {sheet_name}")
        print(f"Header row: {sheet_data['header_row']}")
        print(f"Number of columns: {len(headers)}")
        print(f"\nCOLUMN HEADERS:")
        for idx, header in enumerate(headers, 1):
            print(f"  {idx}. {header}")

        # Track columns
        for header in headers:
            all_columns[header]['count'] += 1
            all_columns[header]['files'].append(file_name)
            if all_columns[header]['example_file'] is None:
                all_columns[header]['example_file'] = file_name

        # Show sample data
        if sample_data:
            print(f"\nSAMPLE DATA (first {len(sample_data)} rows):")
            for row_idx, row in enumerate(sample_data, 1):
                print(f"  Row {row_idx}:")
                for col_idx, val in enumerate(row):
                    if val:
                        print(f"    {headers[col_idx]}: {val}")

# Summary section
print("\n\n")
print("=" * 100)
print("COMPREHENSIVE SUMMARY: ALL UNIQUE COLUMNS ACROSS ALL FILES")
print("=" * 100)

# Sort columns by frequency (most common first)
sorted_columns = sorted(all_columns.items(), key=lambda x: x[1]['count'], reverse=True)

print(f"\nTotal unique column headers found: {len(sorted_columns)}")
print(f"Total files analyzed: {len([f for f in file_formats if f not in ['error']])}")

print("\n" + "=" * 100)
print("COLUMN FREQUENCY TABLE")
print("=" * 100)
print(f"{'Column Name':<50} {'Frequency':<15} {'Example File'}")
print("-" * 100)

for col_name, col_data in sorted_columns:
    print(f"{col_name:<50} {col_data['count']:<15} {col_data['example_file']}")

# Identify potential column name variations
print("\n\n")
print("=" * 100)
print("POTENTIAL COLUMN NAME VARIATIONS (Similar Names)")
print("=" * 100)

# Group similar column names
def normalize_column_name(name):
    """Normalize column name for comparison"""
    return name.lower().replace(' ', '').replace('_', '').replace('-', '')

normalized_groups = defaultdict(list)
for col_name in sorted_columns:
    normalized = normalize_column_name(col_name[0])
    normalized_groups[normalized].append(col_name[0])

# Print groups with variations
print("\nGroups of potentially similar columns:")
for normalized, variations in sorted(normalized_groups.items()):
    if len(variations) > 1:
        print(f"\n  Normalized: '{normalized}'")
        for var in variations:
            count = all_columns[var]['count']
            print(f"    - '{var}' (used in {count} files)")

# Identify different format types
print("\n\n")
print("=" * 100)
print("DISTINCT FILE FORMAT TYPES")
print("=" * 100)

format_signatures = defaultdict(list)
for file_name, sheets in file_formats.items():
    if 'error' in sheets:
        continue
    for sheet_name, sheet_data in sheets.items():
        signature = tuple(sorted(sheet_data['headers']))
        format_signatures[signature].append((file_name, sheet_name))

print(f"\nFound {len(format_signatures)} distinct column combinations:")

for idx, (signature, files) in enumerate(sorted(format_signatures.items(), key=lambda x: len(x[1]), reverse=True), 1):
    print(f"\n{'=' * 100}")
    print(f"FORMAT TYPE #{idx} - Used in {len(files)} sheet(s)")
    print(f"{'=' * 100}")
    print(f"Columns ({len(signature)}):")
    for col in signature:
        print(f"  - {col}")
    print(f"\nFiles using this format:")
    for file_name, sheet_name in files[:5]:  # Show first 5 files
        print(f"  - {file_name} (sheet: {sheet_name})")
    if len(files) > 5:
        print(f"  ... and {len(files) - 5} more files")

print("\n" + "=" * 100)
print("ANALYSIS COMPLETE")
print("=" * 100)
