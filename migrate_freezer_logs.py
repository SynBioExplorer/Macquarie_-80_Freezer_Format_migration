#!/usr/bin/env python3
"""
Freezer Log Migration Script
=============================
Migrates data from old freezer log format to new standardized format.

Author: Claude Code
Date: 2025-10-30
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import re
from datetime import datetime
import logging
from typing import Dict, List, Tuple, Optional
import json
from collections import defaultdict
from multiprocessing import Pool, Manager, cpu_count
import sys
from tqdm import tqdm

# Import path parser
from freezer_path_parser import map_old_sheet_to_new_structure

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('migration.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def process_file_wrapper(args):
    """
    Wrapper function for parallel processing of individual files.
    This function is called by each worker process.

    Args:
        args: Tuple of (file_path, migrator_params)

    Returns:
        Tuple of (migrated_sheets_list, stats_dict, issues_list)
    """
    file_path, migrator = args

    # Create a temporary migrator instance for this worker
    # (each worker needs its own instance to avoid conflicts)
    worker_migrator = FreezerLogMigrator(
        migrator.old_format_dir,
        migrator.new_format_dir,
        migrator.output_dir
    )

    try:
        logger.info(f"Processing file: {file_path.name}")

        # Process the file
        migrated_sheets_dict = worker_migrator.process_old_format_file(file_path)

        # Convert to list format with file path
        migrated_sheets_list = [
            (sheet_name, sheet_df, str(file_path))
            for sheet_name, sheet_df in migrated_sheets_dict.items()
        ]

        # Return results
        return (
            migrated_sheets_list,
            worker_migrator.migration_stats,
            worker_migrator.validation_issues
        )

    except Exception as e:
        logger.error(f"Error processing file {file_path.name}: {str(e)}")
        return ([], {'errors': 1}, [{
            'level': 'ERROR',
            'type': 'File processing error',
            'message': str(e),
            'file': str(file_path),
            'sheet': None,
            'row': None
        }])


class FreezerLogMigrator:
    """Migrates old freezer logs to new standardized format."""

    def __init__(self, old_format_dir: Path, new_format_dir: Path, output_dir: Path):
        self.old_format_dir = Path(old_format_dir)
        self.new_format_dir = Path(new_format_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        # Validation tracking
        self.validation_issues = []
        self.migration_stats = {
            'files_processed': 0,
            'sheets_processed': 0,
            'rows_migrated': 0,
            'rows_skipped': 0,
            'errors': 0,
            'warnings': 0
        }

        # Lab group standardization mapping
        self.lab_group_mapping = {
            'ian paulsen': 'Paulsen',
            'paulsen': 'Paulsen',
            'paulsen group': 'Paulsen',
            'cain': 'Cain',
            'jaschke lab group': 'Jaschke',
            'jaschke lab': 'Jaschke',
            'jaschke group': 'Jaschke',
            "briardo's group": 'Llorente',
            "briardo's gruop": 'Llorente',  # typo
            'briardo llorente': 'Llorente',  # Full name with space
            'williams': 'Williams',
            'tom williams': 'Williams',
            'agf': 'AGF',
            'packer': 'Packer'
        }

        # Invalid lab group entries (should move to comments)
        self.invalid_lab_groups = {
            "pcr didn't work",
            "unsure if pcr is ok",
            "pcr ok",
            "sample thrown away",
            "waiting for sequencing"
        }

        # New format column structure (23 columns total)
        self.new_format_columns = [
            'FREEZER LOCATION',
            'SHELF',
            'Rack Number',
            'Box position',
            'Box Row',
            'Box Column',
            'Label on Box',
            'Sample lodged by',
            'Date lodged',
            'Box description',
            'Contents',
            'Risk Group',
            'NLRD',
            # New columns from old format
            'Organism',
            'Strain',
            'Vector',
            'Marker',
            'Biosafety #',
            'GMO? Yes/No',
            'Sample information',
            'Project information',
            'Sequencing',
            'Comments'
        ]

    def standardize_lab_group(self, value: str, file_path: Optional[str] = None,
                             sheet_name: Optional[str] = None, row_num: Optional[int] = None) -> Tuple[Optional[str], Optional[str]]:
        """
        Standardize lab group value.

        Args:
            value: Lab group value to standardize
            file_path: Source file path for logging (optional)
            sheet_name: Source sheet name for logging (optional)
            row_num: Source row number for logging (optional)

        Returns:
            Tuple of (standardized_value, comment_to_add)
        """
        if pd.isna(value) or value == '':
            return None, None

        value_lower = str(value).strip().lower()

        # Check if it's an invalid entry that should go to comments
        if value_lower in self.invalid_lab_groups:
            return None, f"[From Lab group field: {value}]"

        # Apply standardization mapping
        if value_lower in self.lab_group_mapping:
            return self.lab_group_mapping[value_lower], None

        # If not in mapping, log warning and return as-is
        self.log_validation_issue(
            'WARNING',
            'Unknown lab group',
            f"Lab group value '{value}' not in standardization mapping",
            file_path, sheet_name, row_num
        )
        return str(value).strip(), None

    def parse_tube_position(self, tube_value: str, file_path: Optional[str] = None,
                           sheet_name: Optional[str] = None, row_num: Optional[int] = None) -> Tuple[Optional[str], Optional[int]]:
        """
        Parse tube position into Box Row and Box Column.

        Args:
            tube_value: Tube position value to parse
            file_path: Source file path for logging (optional)
            sheet_name: Source sheet name for logging (optional)
            row_num: Source row number for logging (optional)

        Examples:
            "B5" -> ("B", 5)
            "A1" -> ("A", 1)
            "H12" -> ("H", 12)
        """
        if pd.isna(tube_value) or tube_value == '':
            return None, None

        tube_str = str(tube_value).strip().upper()

        # Try to match pattern like "B5", "A1", "H12"
        match = re.match(r'^([A-H])(\d{1,2})$', tube_str)
        if match:
            return match.group(1), int(match.group(2))

        # Log warning if format doesn't match
        self.log_validation_issue(
            'WARNING',
            'Invalid tube format',
            f"Tube value '{tube_value}' doesn't match expected format (e.g., A1, B5)",
            file_path, sheet_name, row_num
        )
        return None, None

    def standardize_date(self, date_value, file_path: Optional[str] = None,
                        sheet_name: Optional[str] = None, row_num: Optional[int] = None) -> Optional[str]:
        """
        Standardize date to dd/mm/yy format.

        Args:
            date_value: Date value to standardize
            file_path: Source file path for logging (optional)
            sheet_name: Source sheet name for logging (optional)
            row_num: Source row number for logging (optional)
        """
        if pd.isna(date_value) or date_value == '':
            return None

        # If already a datetime object
        if isinstance(date_value, datetime):
            return date_value.strftime('%d/%m/%y')

        # Try to parse various date formats
        date_str = str(date_value).strip()

        # Skip if it's just whitespace
        if not date_str or date_str.isspace():
            return None

        # Common formats to try
        formats = [
            '%d/%m/%Y', '%d/%m/%y',
            '%Y-%m-%d', '%y-%m-%d',
            '%d-%m-%Y', '%d-%m-%y',
            '%d.%m.%Y', '%d.%m.%y'
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%d/%m/%y')
            except ValueError:
                continue

        # If none matched, log warning
        self.log_validation_issue(
            'WARNING',
            'Invalid date format',
            f"Date value '{date_value}' could not be parsed",
            file_path, sheet_name, row_num
        )
        return date_str

    def standardize_gmo_status(self, value: str) -> Optional[str]:
        """Standardize GMO status to 'Yes' or 'No'."""
        if pd.isna(value) or value == '':
            return None

        value_lower = str(value).strip().lower()

        if value_lower in ['yes', 'y', 'true', '1']:
            return 'Yes'
        elif value_lower in ['no', 'n', 'false', '0']:
            return 'No'
        else:
            return str(value).strip()

    def merge_comments(self, existing_comment: str, new_comment: str) -> str:
        """Merge multiple comments together."""
        comments = []

        if existing_comment and str(existing_comment).strip():
            comments.append(str(existing_comment).strip())

        if new_comment and str(new_comment).strip():
            comments.append(str(new_comment).strip())

        return ' | '.join(comments) if comments else ''

    def map_old_to_new_columns(self, old_df: pd.DataFrame, file_path: str, sheet_name: str) -> pd.DataFrame:
        """
        Map old format columns to new format columns.

        Args:
            old_df: DataFrame with old format data
            file_path: Source file path for logging
            sheet_name: Source sheet name for logging

        Returns:
            DataFrame with new format columns
        """
        # Create empty DataFrame with new format columns
        new_df = pd.DataFrame(columns=self.new_format_columns)

        # Normalize old column names (lowercase, remove extra spaces)
        # Handle both string and numeric column names
        old_columns = {col: str(col).strip() if not isinstance(col, (int, float)) else str(col) for col in old_df.columns}
        old_columns_lower = {col.lower(): col for col in old_columns.values()}

        def get_old_column(name_variations: List[str]) -> Optional[str]:
            """Find column in old format by checking variations."""
            for var in name_variations:
                if var.lower() in old_columns_lower:
                    return old_columns_lower[var.lower()]
            return None

        # Map each row
        for idx, old_row in old_df.iterrows():
            new_row = {}

            # 1. FREEZER LOCATION - hardcode to 6WW 326 PC2 Facility
            new_row['FREEZER LOCATION'] = '6WW 326 PC2 Facility'

            # 2. SHELF - extract from file path or existing column
            shelf_col = get_old_column(['Shelf', 'shelf'])
            if shelf_col:
                new_row['SHELF'] = old_row.get(shelf_col)
            else:
                # Try to extract from file path or sheet name
                # Will be filled in by caller
                new_row['SHELF'] = None

            # 3. Rack Number
            rack_col = get_old_column(['Rack Number', 'Rack number', 'rack number', 'Rack/Shelf'])
            if rack_col:
                new_row['Rack Number'] = old_row.get(rack_col)

            # 4. Box position
            box_pos_col = get_old_column(['Box position', 'Box location', 'box position'])
            if box_pos_col:
                new_row['Box position'] = old_row.get(box_pos_col)

            # 5 & 6. Box Row and Box Column - from Tube column
            tube_col = get_old_column(['Tube', 'tube', 'Tube number', 'Tube label'])

            # Handle different tube formats:
            # Format 1: Tube column has combined value like "B5" -> split into Row and Column
            # Format 2: Tube column has just row letter "B", and next column has number "5"

            if tube_col and not pd.isna(old_row.get(tube_col)):
                tube_raw = old_row.get(tube_col)
                # Handle both string and numeric types
                tube_value = str(tube_raw).strip() if tube_raw is not None else ''

                # Skip if empty
                if not tube_value:
                    pass
                # Check if it's combined format like "B5"
                elif len(tube_value) > 1 and tube_value[0].isalpha() and tube_value[1:].isdigit():
                    box_row, box_col = self.parse_tube_position(tube_value, file_path, sheet_name, idx + 1)
                    new_row['Box Row'] = box_row
                    new_row['Box Column'] = box_col
                # Check if it's just a single letter (row only)
                elif len(tube_value) == 1 and tube_value.isalpha():
                    new_row['Box Row'] = tube_value.upper()

                    # Look for column number in next column (often unnamed)
                    tube_col_idx = list(old_columns.values()).index(tube_col)
                    if tube_col_idx + 1 < len(old_df.columns):
                        next_col = old_df.columns[tube_col_idx + 1]
                        next_val = old_row.get(next_col)
                        if not pd.isna(next_val) and str(next_val).strip().isdigit():
                            new_row['Box Column'] = int(next_val)
                else:
                    # Try to parse as is
                    box_row, box_col = self.parse_tube_position(tube_value, file_path, sheet_name, idx + 1)
                    new_row['Box Row'] = box_row
                    new_row['Box Column'] = box_col

            # Also try direct Box Row / Box Column columns if Tube didn't work
            if not new_row.get('Box Row'):
                box_row_col = get_old_column(['Box Row', 'box row'])
                new_row['Box Row'] = old_row.get(box_row_col) if box_row_col else None

            if not new_row.get('Box Column'):
                box_col_col = get_old_column(['Box Column', 'box column'])
                new_row['Box Column'] = old_row.get(box_col_col) if box_col_col else None

            # 7. Label on Box
            label_col = get_old_column(['Label on Box', 'Label on box', 'label on box', 'Lid label'])
            if label_col:
                new_row['Label on Box'] = old_row.get(label_col)

            # 8. Sample lodged by - maps from Owner or Sample lodged by
            owner_col = get_old_column(['Owner', 'owner', 'OWNER'])
            lodged_col = get_old_column(['Sample lodged by', 'sample lodged by'])

            if owner_col and not pd.isna(old_row.get(owner_col)):
                new_row['Sample lodged by'] = old_row.get(owner_col)
            elif lodged_col:
                new_row['Sample lodged by'] = old_row.get(lodged_col)

            # 9. Date lodged
            date_col = get_old_column(['Date lodged (dd/mm/yy)', 'Date lodged', 'Date', 'Date of congelation'])
            if date_col:
                new_row['Date lodged'] = self.standardize_date(old_row.get(date_col), file_path, sheet_name, idx + 1)

            # 10. Box description
            box_desc_col = get_old_column(['Box description', 'box description'])
            if box_desc_col:
                new_row['Box description'] = old_row.get(box_desc_col)

            # 11. Contents
            contents_col = get_old_column(['Contents', 'contents', 'Contain'])
            if contents_col:
                new_row['Contents'] = old_row.get(contents_col)

            # 12. Risk Group
            risk_col = get_old_column(['Risk Group', 'RiskGroup', 'RG', 'risk group'])
            if risk_col:
                new_row['Risk Group'] = old_row.get(risk_col)

            # 13. NLRD
            nlrd_col = get_old_column(['NLRD', 'nlrd'])
            if nlrd_col:
                new_row['NLRD'] = old_row.get(nlrd_col)

            # 14. Organism
            organism_col = get_old_column(['Organism', 'organism'])
            if organism_col:
                new_row['Organism'] = old_row.get(organism_col)

            # 15. Strain
            strain_col = get_old_column(['Strain', 'strain'])
            if strain_col:
                new_row['Strain'] = old_row.get(strain_col)

            # 16. Vector - merge from Vector and Plasmid columns
            vector_col = get_old_column(['Vector', 'vector'])
            plasmid_col = get_old_column(['Plasmid', 'plasmid'])

            vector_value = old_row.get(vector_col) if vector_col else None
            plasmid_value = old_row.get(plasmid_col) if plasmid_col else None

            if not pd.isna(vector_value) and vector_value:
                new_row['Vector'] = vector_value
            elif not pd.isna(plasmid_value) and plasmid_value:
                new_row['Vector'] = plasmid_value

            # 17. Marker
            marker_col = get_old_column(['Marker', 'marker', 'Marker yeast', 'Marker bacteria', 'Antibiotic Marker'])
            if marker_col:
                new_row['Marker'] = old_row.get(marker_col)

            # 18. Biosafety #
            biosafety_col = get_old_column(['Biosafety #', 'biosafety #', 'Biosafety pending'])
            if biosafety_col:
                new_row['Biosafety #'] = old_row.get(biosafety_col)

            # 19. GMO? Yes/No
            gmo_col = get_old_column(['GMO? Yes/No', 'Contains GMOs? (Yes/No)', 'gmo? yes/no'])
            if gmo_col:
                new_row['GMO? Yes/No'] = self.standardize_gmo_status(old_row.get(gmo_col))

            # 20. Sample information
            sample_info_col = get_old_column(['Sample information', 'sample information', 'Sample info'])
            if sample_info_col:
                new_row['Sample information'] = old_row.get(sample_info_col)

            # 21. Project information
            proj_col = get_old_column(['Project information', 'project information', 'Project info',
                                       'Project information/origin/plasmid from'])
            if proj_col:
                new_row['Project information'] = old_row.get(proj_col)

            # 22. Sequencing
            seq_col = get_old_column(['Sequencing', 'sequencing', 'Genotyping|Sequencing',
                                      'Sequencing result', 'Date of sequencing (mini-prep + sequencing done from the same culture used for glycerol stock)'])
            if seq_col:
                new_row['Sequencing'] = old_row.get(seq_col)

            # 23. Comments - handle lab group invalids and merge multiple comment sources
            comments = []

            # Get existing comments
            comment_col = get_old_column(['Comments', 'comments', 'Comment', 'Coments',
                                          'Additional comments', 'Additional comment', 'Note'])
            if comment_col and not pd.isna(old_row.get(comment_col)):
                comments.append(str(old_row.get(comment_col)))

            # Handle Lab group standardization and invalid entries
            lab_group_col = get_old_column(['Lab group', 'Lab Group', 'lab group'])
            if lab_group_col and not pd.isna(old_row.get(lab_group_col)):
                standardized_lab, comment_from_lab = self.standardize_lab_group(
                    old_row.get(lab_group_col),
                    file_path,
                    sheet_name,
                    idx + 1  # Add 1 for 1-based row numbering
                )

                # If invalid lab group entry, add to comments
                if comment_from_lab:
                    comments.append(comment_from_lab)

                # Store standardized lab group in a temporary column for reference
                # (Note: Lab group is not in new format columns, but we track it)
                new_row['_Lab_Group'] = standardized_lab

            # Merge all comments
            new_row['Comments'] = ' | '.join(comments) if comments else None

            # Add row to new dataframe
            # Use loc to avoid FutureWarning about concatenation
            row_idx = len(new_df)
            for col, val in new_row.items():
                new_df.loc[row_idx, col] = val

        return new_df

    def log_validation_issue(self, level: str, issue_type: str, message: str,
                            file_path: Optional[str], sheet_name: Optional[str],
                            row_num: Optional[int]):
        """Log validation issues for reporting."""
        issue = {
            'level': level,
            'type': issue_type,
            'message': message,
            'file': file_path,
            'sheet': sheet_name,
            'row': row_num,
            'timestamp': datetime.now().isoformat()
        }
        self.validation_issues.append(issue)

        # Also log to logger
        log_msg = f"{issue_type}: {message}"
        if file_path:
            log_msg += f" [File: {file_path}"
            if sheet_name:
                log_msg += f", Sheet: {sheet_name}"
            if row_num:
                log_msg += f", Row: {row_num}"
            log_msg += "]"

        if level == 'ERROR':
            logger.error(log_msg)
        elif level == 'WARNING':
            logger.warning(log_msg)
        else:
            logger.info(log_msg)

    def process_old_format_file(self, file_path: Path) -> Dict[str, pd.DataFrame]:
        """
        Process a single old format Excel file.

        Returns:
            Dict mapping sheet names to DataFrames with migrated data
        """
        logger.info(f"Processing file: {file_path.name}")

        try:
            # Load Excel file
            xl_file = pd.ExcelFile(file_path)
            migrated_sheets = {}

            for sheet_name in xl_file.sheet_names:
                logger.info(f"  Processing sheet: {sheet_name}")

                try:
                    # Read sheet with a maximum row limit to handle sheets with excessive formatting
                    # Use nrows parameter to limit reading (skip sheets with millions of formatted rows)
                    df = pd.read_excel(xl_file, sheet_name=sheet_name, nrows=10000)

                    # Skip empty sheets
                    if df.empty or len(df) == 0:
                        logger.info(f"    Skipping empty sheet: {sheet_name}")
                        continue

                    # Drop rows that are completely empty
                    df = df.dropna(how='all')

                    # Check if first row contains warning message (skip it)
                    if not df.empty and len(df) > 0:
                        first_cell = str(df.iloc[0, 0]).lower()
                        if 'remember to fill' in first_cell or 'warning' in first_cell:
                            df = df.iloc[2:].reset_index(drop=True)  # Skip first 2 rows
                            # Reset column headers
                            df.columns = df.iloc[0]
                            df = df.iloc[1:].reset_index(drop=True)

                    # Map to new format
                    migrated_df = self.map_old_to_new_columns(df, str(file_path), sheet_name)

                    if not migrated_df.empty:
                        migrated_sheets[sheet_name] = migrated_df
                        self.migration_stats['sheets_processed'] += 1
                        self.migration_stats['rows_migrated'] += len(migrated_df)

                except Exception as e:
                    logger.error(f"    Error processing sheet {sheet_name}: {str(e)}")
                    self.log_validation_issue('ERROR', 'Sheet processing error',
                                             str(e), str(file_path), sheet_name, None)
                    self.migration_stats['errors'] += 1

            self.migration_stats['files_processed'] += 1
            return migrated_sheets

        except Exception as e:
            logger.error(f"Error processing file {file_path.name}: {str(e)}")
            self.log_validation_issue('ERROR', 'File processing error',
                                     str(e), str(file_path), None, None)
            self.migration_stats['errors'] += 1
            return {}

    def create_new_format_excel(self, freezer_num: int, shelf: str,
                               rack_data: Dict[str, pd.DataFrame]) -> Path:
        """
        Create a new format Excel file with multiple sheets for each rack.

        Args:
            freezer_num: Freezer number (1, 2, or 3)
            shelf: Shelf letter (A, B, C, D, E)
            rack_data: Dict mapping rack numbers to DataFrames

        Returns:
            Path to created Excel file
        """
        filename = f"6WW 326 PC2 Facility - Freezer {freezer_num} Shelf {shelf}.xlsx"
        output_path = self.output_dir / filename

        logger.info(f"Creating new format file: {filename}")

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add sheets for each rack
        for rack_num in sorted(rack_data.keys()):
            df = rack_data[rack_num]

            # Extract numeric rack number from rack identifier (e.g., "A1" -> "1", "B2" -> "2")
            rack_number = ''.join(filter(str.isdigit, rack_num))
            if not rack_number:
                rack_number = rack_num  # fallback to full rack identifier

            # Create sheet name matching new format: "Shelf A Rack 1_Empty"
            sheet_name = f"Shelf {shelf} Rack {rack_number}_Empty"
            ws = wb.create_sheet(sheet_name)

            # Add warning row
            ws['A1'] = "REMEMBER TO FILL IN RELEVANT CELLS FOR RISK GROUP AND/ OR NLRD DETAILS IF APPLICABLE FOR YOUR SAMPLES"
            ws['A1'].font = Font(bold=True, color="FF0000")

            # Add empty row
            # Row 2 is empty

            # Add headers in row 3
            for col_idx, col_name in enumerate(self.new_format_columns, start=1):
                cell = ws.cell(row=3, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Add data starting from row 4
            for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
                for col_idx, col_name in enumerate(self.new_format_columns, start=1):
                    value = row.get(col_name)
                    if pd.notna(value) and value != '':
                        ws.cell(row=row_idx, column=col_idx, value=value)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Saved: {output_path}")

        return output_path

    def generate_validation_report(self):
        """Generate comprehensive validation report."""
        report_path = self.output_dir / 'validation_report.txt'

        with open(report_path, 'w') as f:
            f.write("="*80 + "\n")
            f.write("FREEZER LOG MIGRATION VALIDATION REPORT\n")
            f.write("="*80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            # Migration statistics
            f.write("MIGRATION STATISTICS\n")
            f.write("-"*80 + "\n")
            for key, value in self.migration_stats.items():
                f.write(f"{key.replace('_', ' ').title()}: {value}\n")
            f.write("\n")

            # Validation issues by level
            errors = [i for i in self.validation_issues if i['level'] == 'ERROR']
            warnings = [i for i in self.validation_issues if i['level'] == 'WARNING']
            infos = [i for i in self.validation_issues if i['level'] == 'INFO']

            f.write(f"Total Issues: {len(self.validation_issues)}\n")
            f.write(f"  - Errors: {len(errors)}\n")
            f.write(f"  - Warnings: {len(warnings)}\n")
            f.write(f"  - Info: {len(infos)}\n\n")

            # List errors
            if errors:
                f.write("ERRORS\n")
                f.write("-"*80 + "\n")
                for issue in errors:
                    f.write(f"[{issue['type']}] {issue['message']}\n")
                    if issue['file']:
                        f.write(f"  File: {issue['file']}\n")
                        if issue['sheet']:
                            f.write(f"  Sheet: {issue['sheet']}\n")
                        if issue['row']:
                            f.write(f"  Row: {issue['row']}\n")
                    f.write("\n")

            # List warnings - grouped by type for better readability
            if warnings:
                f.write("WARNINGS\n")
                f.write("-"*80 + "\n")

                # Group warnings by type
                warnings_by_type = defaultdict(list)
                for issue in warnings:
                    warnings_by_type[issue['type']].append(issue)

                # Display each warning type with all instances
                for warning_type in sorted(warnings_by_type.keys()):
                    issues = warnings_by_type[warning_type]
                    f.write(f"\n{warning_type} ({len(issues)} occurrences)\n")
                    f.write("-"*40 + "\n")

                    # Show ALL instances with full details
                    for issue in issues:
                        f.write(f"  {issue['message']}\n")
                        if issue['file']:
                            file_name = Path(issue['file']).name
                            f.write(f"    File: {file_name}\n")
                        if issue['sheet']:
                            f.write(f"    Sheet: {issue['sheet']}\n")
                        if issue['row'] is not None:
                            f.write(f"    Row: {issue['row']}\n")
                        f.write("\n")

                f.write("\n")

        logger.info(f"Validation report saved to: {report_path}")

        # Also save issues as JSON for programmatic access
        json_path = self.output_dir / 'validation_issues.json'
        with open(json_path, 'w') as f:
            json.dump(self.validation_issues, f, indent=2)
        logger.info(f"Validation issues JSON saved to: {json_path}")

    def migrate_all(self, num_processes=10):
        """Main migration process with parallel processing.

        Args:
            num_processes: Number of parallel processes to use (default: 10)
        """
        logger.info("="*80)
        logger.info("Starting Freezer Log Migration (Parallel Processing)")
        logger.info(f"Using {num_processes} parallel processes")
        logger.info("="*80)

        # Data structure: freezer_data[freezer_num][shelf][rack] = DataFrame
        freezer_data = defaultdict(lambda: defaultdict(lambda: defaultdict(pd.DataFrame)))

        # Process all old format files
        all_files = list(self.old_format_dir.rglob('*.xlsx'))
        logger.info(f"Found {len(all_files)} Excel files to process")

        # Filter out template and allocation files
        files_to_process = []
        for file_path in all_files:
            # Skip template files
            if 'BLANK TEMPLATE' in file_path.name or 'template' in file_path.name.lower():
                logger.info(f"Skipping template file: {file_path.name}")
                continue

            # Skip allocation files
            if 'allocation' in file_path.name.lower():
                logger.info(f"Skipping allocation file: {file_path.name}")
                continue

            files_to_process.append(file_path)

        logger.info(f"Processing {len(files_to_process)} files in parallel...")

        # Process files in parallel using multiprocessing with progress bar
        results = []
        with Pool(processes=num_processes) as pool:
            with tqdm(total=len(files_to_process), desc="Migrating files", unit="file") as pbar:
                for result in pool.imap_unordered(process_file_wrapper, [(file_path, self) for file_path in files_to_process]):
                    results.append(result)
                    pbar.update(1)

        # Merge results from all workers
        logger.info("Merging results from parallel workers...")
        for file_results in results:
            if file_results is None:
                continue

            migrated_sheets, file_stats, file_issues = file_results

            # Update stats
            self.migration_stats['files_processed'] += file_stats.get('files_processed', 0)
            self.migration_stats['sheets_processed'] += file_stats.get('sheets_processed', 0)
            self.migration_stats['rows_migrated'] += file_stats.get('rows_migrated', 0)
            self.migration_stats['errors'] += file_stats.get('errors', 0)
            self.migration_stats['warnings'] += file_stats.get('warnings', 0)

            # Merge validation issues
            self.validation_issues.extend(file_issues)

            # Organize migrated sheets by freezer/shelf/rack
            for sheet_name, sheet_df, file_path_str in migrated_sheets:
                # Parse freezer, shelf, rack from file path and sheet name
                file_path = Path(file_path_str)
                freezer_num, shelf, rack = map_old_sheet_to_new_structure(file_path, sheet_name)

                if not freezer_num:
                    self.log_validation_issue(
                        'WARNING',
                        'Missing freezer number',
                        f"Could not determine freezer number for {file_path.name} - {sheet_name}",
                        str(file_path), sheet_name, None
                    )
                    continue

                if not shelf:
                    self.log_validation_issue(
                        'WARNING',
                        'Missing shelf',
                        f"Could not determine shelf for {file_path.name} - {sheet_name}",
                        str(file_path), sheet_name, None
                    )
                    # If we have a rack, we can extract shelf from it (e.g., "A1" -> "A")
                    if rack and len(rack) >= 1 and rack[0].isalpha():
                        shelf = rack[0]
                        logger.info(f"  Extracted shelf '{shelf}' from rack '{rack}'")

                if not rack:
                    self.log_validation_issue(
                        'WARNING',
                        'Missing rack',
                        f"Could not determine rack for {file_path.name} - {sheet_name}",
                        str(file_path), sheet_name, None
                    )
                    continue

                # Fill in SHELF and Rack Number columns
                sheet_df['SHELF'] = shelf
                sheet_df['Rack Number'] = rack

                # Append to or create DataFrame for this freezer/shelf/rack
                if freezer_data[freezer_num][shelf][rack].empty:
                    freezer_data[freezer_num][shelf][rack] = sheet_df
                else:
                    # Concatenate if there are multiple sheets for same rack
                    freezer_data[freezer_num][shelf][rack] = pd.concat(
                        [freezer_data[freezer_num][shelf][rack], sheet_df],
                        ignore_index=True
                    )

                logger.info(f"  Mapped to: Freezer {freezer_num}, Shelf {shelf}, Rack {rack}")

        # Create new format Excel files organized by freezer and shelf
        logger.info("\nCreating new format Excel files...")

        for freezer_num in sorted(freezer_data.keys()):
            # Filter out None values from shelves before sorting
            valid_shelves = [s for s in freezer_data[freezer_num].keys() if s is not None]
            for shelf in sorted(valid_shelves):
                rack_data_for_shelf = freezer_data[freezer_num][shelf]

                if rack_data_for_shelf:
                    # Create new format Excel file
                    output_path = self.create_new_format_excel(
                        freezer_num,
                        shelf,
                        rack_data_for_shelf
                    )
                    logger.info(f"Created: {output_path.name}")

        # Generate validation report
        self.generate_validation_report()

        logger.info("="*80)
        logger.info("Migration Complete")
        logger.info("="*80)
        logger.info(f"Files processed: {self.migration_stats['files_processed']}")
        logger.info(f"Rows migrated: {self.migration_stats['rows_migrated']}")
        logger.info(f"Errors: {self.migration_stats['errors']}")
        logger.info(f"Warnings: {self.migration_stats['warnings']}")


def main():
    """Main entry point."""
    import sys

    # Paths
    base_dir = Path("/Users/felix/Library/CloudStorage/OneDrive-SharedLibraries-MacquarieUniversity/Australian Genome Foundry - AWS cloud infrastructure/09_Freezer_Log_formatting")
    old_format_dir = base_dir / "older freezer log format"
    new_format_dir = base_dir / "new freezer log format"
    output_dir = base_dir / "migrated_logs"

    # Create migrator
    migrator = FreezerLogMigrator(old_format_dir, new_format_dir, output_dir)

    # Run migration with 10 parallel processes
    try:
        migrator.migrate_all(num_processes=10)
        print("\n✓ Migration completed successfully!")
        print(f"✓ Output directory: {output_dir}")
        print(f"✓ Check validation_report.txt for details")
    except Exception as e:
        logger.error(f"Migration failed: {str(e)}", exc_info=True)
        sys.exit(1)


if __name__ == '__main__':
    main()
